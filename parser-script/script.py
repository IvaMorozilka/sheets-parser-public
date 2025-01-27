from utils import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import argparse
import os
from icecream import ic
ic.disable()

def transform_pipeline(sheet, input, output, modded):

    # Проверка наличия активного листа
    if sheet is None:
        raise Exception("Файл не содержит активного листа.")

    if sheet.max_row == 0 or sheet.max_column == 0:
        raise Exception("Активный лист пустой.")
    
    if sheet['A1'].value is None or sheet['A1'].value == "":
        raise Exception('Данные должны начинаться с ячейки A1. ')

    # Разъеденим все объедененные ячейки, они мешают.
    sheet = unmerge_cells(sheet)

    # Удаление столбца A
    process_and_delete_column(sheet, "A")

    # Удаление всех фильтров
    sheet.auto_filter.ref = None

    # Раскрываем скрытые строки и столбцы
    sheet = remove_hidden_cells(sheet)

    # Увеличим ширину столбцов, чтобы влезали цифры и было красиво
    set_column_width(sheet, ["B", "C", "D", "E", "F", "G", "H", "I"], 18)
    set_column_width(sheet, ["A"], 63)

    # Ищем индекс ячейки в первой строке с "итого"
    for cell in next(sheet.iter_rows()):
        if "Итого" in cell.value:
            idx = cell.col_idx
            break

    process_header(sheet)
    sheet = replace_bad_values(sheet)

    otvetst_col_letter = get_column_letter(sheet.max_column - 1)

    # Этап добавления рaсчетов
    # data[0] расчеты в руб, data[1] - в %
    data, logs = calculate_additional_data(sheet)

    for log in logs:
        if any(value == "" for value in log.values()):
            raise Exception(
                'При расчетах произошла ошибка. Не были найдены ячейки "Развитие" или "Сопровождение" в колонке "Наименование показателя эффективности и результативности деятельности учреждения". Возможно опечатки.'
            )

    last_row = find_last_row_with_word(sheet, otvetst_col_letter, "Бекетова") + 1
    col_rub = get_column_letter(find_column_index_by_header(sheet, ["Итого", "руб"]))
    col_perc = get_column_letter(find_column_index_by_header(sheet, ["Итого", "%"]))

    sheet.insert_rows(last_row, 6)
    for key, value_rub, value_perc in zip(
        data[0].keys(), data[0].values(), data[1].values()
    ):
        sheet[f"A{last_row}"].value = key
        sheet[f"{col_rub}{last_row}"] = value_rub
        sheet[f"{col_rub}{last_row}"].number_format = "0.00"

        sheet[f"{col_perc}{last_row}"] = value_perc
        sheet[f"{col_perc}{last_row}"].number_format = "0.00%"

        last_row += 1

    # Переносим Гречушкина выше
    move_and_replace_rows(sheet, otvetst_col_letter, "Гречушкин", 147)

    # Заполняем id
    sheet.delete_rows(2)
    delete_empty_rows(sheet)
    fill_column_with_ids(sheet, 2, 2, get_column_letter(sheet.max_column))
    apply_borders_to_all_cells(sheet)
    apply_font_to_all_cells(sheet, "Times New Roman", 11)

    try:
        current_dir = os.getcwd()
        if modded:
            dir_path = os.path.join(current_dir, 'modded_files')
            if not os.path.exists(dir_path):
                os.mkdir(dir_path)
            workbook.save(os.path.join(dir_path, "изм_" + input))
            return True, f"Файл успешно обработан и сохранен {os.path.join(dir_path, 'изм_' + input)}."
        else:
            workbook.save(output)
            return True, "Файл успешно обработан и сохранен {output}"
        
    except Exception as e:
        raise Exception(f"Произошла ошибка при сохранении файла. {e}")




if __name__ == "__main__":
    
    # Добавляем аргументы
    parser = argparse.ArgumentParser("Парсер документов Excel. Трансформирует таблицу по заданным правилам.")
    parser.add_argument("input",  help="Путь к входному файлу Excel. Поддерживаемые расширения: xlsx/xlsm/xltx/xltm")
    parser.add_argument("--output", "-o", help="Путь к выходному файлу (обработанному). Если используется без аргумента -m указать название и расширение сохраняемого файла.", default="modified.xlsx")
    parser.add_argument("--modded", "-m", help="Добавить приставку modded_ к входному файлу.", action="store_true")
    parser.add_argument("--verbose", "-v", help="Включить минимальное логирование (пока только расчетов)", action="store_true")

    # Парсим аргументы
    args = parser.parse_args()

    try:
        workbook = load_workbook(args.input, data_only=True, read_only=False)
        sheet = workbook.active
    except Exception as e:
        raise Exception(f"Произошла ошибка при открытии файла. {e}")
    
    if args.verbose:
        ic.enable()
    
    transform_pipeline(sheet, args.input, args.output, args.modded)


